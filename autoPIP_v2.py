# -*- coding: utf-8 -*-
"""
Created on Mon Sep 10 16:03:12 2018

@author: bec
"""
import ee
import os
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from WA_Hyperloop import becgis
import numpy as np  
import gdal
import cartopy.crs as ccrs
from cartopy.mpl.gridliner import LONGITUDE_FORMATTER, LATITUDE_FORMATTER
import cartopy.feature as feature
from mpl_toolkits.axes_grid1 import make_axes_locatable
import cartopy.io.shapereader as shpreader
import urllib
import zipfile
from openpyxl import load_workbook
import scipy.stats as sts
import win32com.client

ee.Initialize()

class trend_sheet(object):
    
    def fixTIMERES(self, VAR):
        
        def _fixTIMERES(year):

            year = ee.Number(year)

            VARnew = VAR.filterDate(ee.Date.fromYMD(year, 1, 1), ee.Date.fromYMD(year.add(1), 1, 1))
            
            time = VARnew.first().get('system:time_start')
            multiplier = VARnew.first().get('multiplier')
            unit = VARnew.first().get('unit')
            
            VARyear = VARnew.reduce(ee.Reducer.sum())
          
            VARyear = VARyear.set('system:time_start', time).set('multiplier', multiplier).set('unit', unit).set('year', year)
            
            return VARyear

        years = ee.List.sequence(self.Start, self.End, 1)
        
        VARyearly = ee.ImageCollection(years.map(_fixTIMERES))
        
        return VARyearly

    def calc_AGBP(self, NPP):
        def _calc_AGBP(image):
              time = image.get('system:time_start')
              multiplier = image.get('multiplier')
              unit = image.get('unit')
              agbp = image.multiply(0.01444).multiply(image.metadata('n_days_extent'))
              agbp = agbp.set('system:time_start', time).set('multiplier', multiplier).set('unit', unit)
              return agbp
        NPP = ee.ImageCollection(NPP)
        
        AGBP = NPP.map(_calc_AGBP)
        
        return AGBP
        
    
    def __init__(self, name, AGBP, ET, P, ETref, CountryShape, OutputFolder, NPP = None):
        
        # Turn of the visibility of matplotlib figures (i.e. figures remain invisble until saved as a file or after plt.show())
        plt.ioff()
        
        self.Name = name
        
        # Create figure object
        self.initFigure()
        
        # Determine start and end years available in ET image collection
        self.Start = ee.Date(ee.ImageCollection(ET).aggregate_min('system:time_start')).get('year').getInfo()
        self.End = ee.Date(ee.ImageCollection(ET).aggregate_max('system:time_start')).get('year').getInfo()
        
        # Calculate AGBP if NPP is given instead of AGBP
        if NPP != None:
            self.AGBP_dekad = self.calc_AGBP(NPP)
            AGBP = self.fixTIMERES(self.AGBP_dekad)
        
        # Create dictionary with all the image collections
        self.VARS = {"ET": ee.ImageCollection(ET).set('variable', 'ET Actual'),
                     "P": ee.ImageCollection(P).set('variable', 'Precipitation'),
                     "AGBP": ee.ImageCollection(AGBP).set('variable', 'Above Ground Biomass Prod.'),
                     "ET0": ee.ImageCollection(ETref).set('variable', "Reference ET")}

        # Adjust the AGBP unit and correct the multiplier value in the metadata
        self.correct_AGBP()
        
        # Determine the resolution to be used
        self.scale = ee.Image(self.VARS['ET'].select('b1_sum').first()).projection().nominalScale()

        # Create a mask image from the shapefile
        self.CountryShape = ee.FeatureCollection(CountryShape)
        self.CountryMask = self.CountryShape.map(lambda sh: sh.set('1', 1)).reduceToImage(['1'], ee.Reducer.max()).unmask().cast({"max": "uint8"})

        # Calculate the total area of the AOI
        self.Area = self.CountryShape.geometry().area().getInfo()
        
        self.OutFldr = OutputFolder
        
        # Set which variables to show in which columns
        self.Columns = {'AGBP': {'pos': 0}, 'ET': {'pos': 1}, 'WP': {'pos': 2}, 'YIELD': {'pos': 0}, 'P': {}, 'ET0': {}}
        
        # Set the Harvest Index and Water Content to be used
        # --> This should become input into the function. If HI = 1.0 and WC = 0.0, the first column should show 
        # --> the AGBP, otherwise the Yield. Also, then the WP needs to be calculated with base = 'YIELD' (also see make2)
        self.HI = 1.00
        self.WC = 0.00
        
        # Convert time resolution to yearly for required datasets
        for varname, VAR in self.VARS.items():
            if VAR.first().get('time_resolution').getInfo() != 'YEAR':
                print(varname)
                self.VARS[varname] = self.fixTIMERES(VAR)
        

    def initFigure(self):
        self.f = plt.figure(1)
        self.f.clf()
        self.f.set_dpi(96)
        self.f.set_size_inches(16.54, 11.69) # A3 size
        self.gs = gridspec.GridSpec(3, 3, height_ratios=[1,1,2])
        
    def correct_AGBP(self):
        def _correct_AGBP(img):
            date = img.get('system:time_start')
            img = img.multiply(10)
            img = img.set('multiplier', 0.1)
            img = img.set('unit', 'kg/ha').set('system:time_start', date).set('time_resolution', 'YEAR')     # Correct the multiplier value (should be 0.1?) and convert unit to kg.ha 
            return img
        self.VARS['AGBP'] = self.VARS['AGBP'].map(_correct_AGBP)
        
    def calcWP(self, base = 'AGBP'):
        
        def _calcWP(year):
            agbp = ee.Image(self.VARS[base].filterMetadata('system:time_start', 'equals', year).first())
            agbp = agbp.multiply(ee.Number(agbp.get("multiplier"))) ## kgC/ha
              
            et = ee.Image(self.VARS['ET'].filterMetadata('system:time_start', 'equals', year).first())
            et = et.multiply(ee.Number(et.get("multiplier"))) ## mm
            wp = agbp.divide(et.multiply(10)).set('system:time_start', year).set('multiplier', 1.0).set('unit', 'kg/m3')
            
            return wp
    
        years = ee.List(self.VARS[base].aggregate_array('system:time_start'))
        self.VARS['WP'] = ee.ImageCollection(years.map(_calcWP)).set('variable', 'Water Productivity')
        
    def calcTREND(self, VAR):

        start = VAR.aggregate_min('system:time_start')
        first = ee.Image(VAR.filterMetadata('system:time_start', 'equals', start).first()).select("b1_sum")
          
        trend = VAR.select("b1_sum").formaTrend()
        
        def add_timeband(VARimg):
            date = ee.Date(VARimg.get("system:time_start"))
            return ee.Image(date.get("year")).cast({"constant": "int32"}).addBands(VARimg)
      
        with_time = VAR.map(add_timeband);
        pearson = with_time.reduce(ee.Reducer.pearsonsCorrelation());
      
        trend = trend.select('long-trend').divide(first).multiply(100)
        trend = trend.addBands(self.CountryMask).addBands(pearson)
        
        return trend
    
    def downloadImage(self, image, output_fh, shape, scale):
    
        region = ee.Geometry(shape.geometry().bounds(1).getInfo()).toGeoJSONString()
        
        params = {
                  'name':'test',
                  'crs': 'EPSG:4326',
                  'scale': scale,
                  'region': region
                 }
        
        url = image.getDownloadURL(params)
    
        succes = True
        while succes:
            try:
                print "start download"
                urllib.urlretrieve(url, output_fh)
                zip_ref = zipfile.ZipFile(output_fh, 'r')
                zip_ref.extractall(output_fh[:-4])
                zip_ref.close()
                succes = False
            except:
                pass    
            
    def createTRENDMAP(self, varname):
        trend = self.calcTREND(self.VARS[varname])
        output_fh = os.path.join(self.OutFldr, '{0}.zip'.format(varname))
        
        self.factor = np.interp(self.Area, [50000, 625000000], [0.1, 1.0])
        scale = self.scale.multiply(self.factor).getInfo()

        self.downloadImage(trend, output_fh, self.CountryShape, scale)
        
        TREND = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.long-trend.tif'))
        MASK = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.max.tif'))
        PVALUE = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.p-value.tif'))
        
        assert np.shape(TREND) == np.shape(MASK), "resolution dont match"
        
        TREND[MASK == 0] = np.nan
        TREND[PVALUE > 0.1] = np.nan

        PVALUE[PVALUE <= 0.1] = np.nan
        PVALUE[MASK == 0] = np.nan
        PVALUE[~np.isnan(PVALUE)] = 1.
        
        AREA = becgis.MapPixelAreakm(os.path.join(output_fh[:-4], 'test.max.tif'))
        self.PixelArea = np.mean(AREA[MASK == 1]) * 100 # ha
        
        self.Columns[varname]['TREND'] = TREND
        self.Columns[varname]['PVALUE'] = PVALUE
        
        ds = gdal.Open(os.path.join(output_fh[:-4], 'test.long-trend.tif'))
        gt = ds.GetGeoTransform()
        n_cols = ds.RasterXSize
        n_rows = ds.RasterYSize
        ds = None
        self.extent_ll = (gt[0], gt[0] + (gt[1] * n_cols), gt[3] + (gt[5] * n_rows), gt[3])

    def createTS(self, varname, reducer = ee.Reducer.mean()):
        def _createTS(VARimg):
            date = ee.Date(VARimg.get("system:time_start"))
                  
            VARyear = VARimg.multiply(ee.Number(VARimg.get("multiplier"))).reduceRegion(reducer = reducer, geometry = self.CountryShape, maxPixels = 1e12, scale = self.scale)
                                                     
            VARft = ee.Feature(None, {"date": date,
                                      "year": date.get("year"),
                                      "unit": VARimg.get("unit")})
                                                
            VARft = VARft.set(VARyear)
                  
            return VARft

        VARts = self.VARS[varname].map(_createTS)
        
        unit = '[' + VARts.aggregate_first('unit').getInfo().replace(u'\xb2', '2') + '/yr]'
        
        VARts = np.array([VARts.aggregate_array('year').getInfo(), VARts.aggregate_array('b1_sum').getInfo()])
        
        if not reducer == ee.Reducer.mean():
            self.Columns[varname]['TS95'] = VARts
            self.Columns[varname]['unit'] = unit
        else:
            self.Columns[varname]['TS'] = VARts
            self.Columns[varname]['unit'] = unit            
    
    def plotTS(self, varname):
        ax = plt.subplot(self.gs[0, self.Columns[varname]['pos']])
            
        yrs = self.Columns[varname]['TS'][0].astype(int)
        vals = self.Columns[varname]['TS'][1]
        
        unit = self.Columns[varname]['unit']
    
        ax.set_ylabel("{0} {1}".format(varname, unit.replace('/yr', '')))
        
        slope, intercept, r_value, p_value, std_err = sts.linregress(yrs, vals)
        
        fit = np.polyfit(yrs, vals, 1)
        fit_fn = np.poly1d(fit) 
        
        yrs_long = [yrs[0]-1, yrs[-1]+1]
    
        ax.plot(yrs_long, fit_fn(yrs_long), 'k--')

        ax.set_ylim([0.93 * np.min(vals), 1.02 * np.max(vals)])
        ax.set_xlim(yrs_long)
        ax.set_title("{4} \n slope = {0:.2f} {1}, p = {2:.3f}, r = {3:.2f}".format(slope, unit[1:-1], p_value, r_value, varname))
        ax.set_facecolor('#d8dcd6')
        ax.grid(b=True, which='Major', color='0.65',linestyle='--', zorder=0)
        ax.bar(yrs, vals, color = '#005c30', zorder = 2)
        ax.set_xticks(yrs)
        ax.set_xticklabels(['`' + yr[-2:] for yr in yrs.astype(str)])
        
    def plotHIST(self, varname):
        ax = plt.subplot(self.gs[1, self.Columns[varname]['pos']])
        ax.set_xlabel("Trend [%/yr]")
        ax.set_ylabel("Area [ha]")
        ax.set_facecolor('#d8dcd6')
        ax.grid(b=True, which='Major', color='0.65',linestyle='--', zorder=0)

        rng = (np.floor(np.nanmin(self.Columns[varname]['TREND']) / 10)*10, np.ceil(np.nanmax(self.Columns[varname]['TREND']) / 10)*10)
        
        try:
            bns = np.min([int(np.diff(rng)[0]), 21])
        except:
            bns = 0
        
        print(varname, bns, rng)
        
        if bns != 0:
            pixels, bins = np.histogram(self.Columns[varname]['TREND'][~np.isnan(self.Columns[varname]['TREND'])], range = rng, bins = bns)
    
            bins = (bins + np.diff(bins)[0]/2)[:-1]
            
            wdt = np.diff(bins)[0]
    
            if varname == "ET":
                colors = ['#005c30','#940121']
            else:
                colors = ['#940121','#005c30']
                                  
            ax.bar(bins[bins < 0], pixels[bins < 0], color = colors[0], width = wdt, zorder = 2)
            ax.bar(bins[bins >= 0], pixels[bins >= 0], color = colors[1], width = wdt, zorder = 2)
            
        new_lbls = [str(int(item * self.PixelArea)) for item in ax.get_yticks().tolist()]
        
        ax.set_yticklabels(new_lbls)
        
    def plotTRENDMAP(self, varname):
        ax = plt.subplot(self.gs[2, self.Columns[varname]['pos']], projection = ccrs.PlateCarree())
        
        crs_lonlat = ccrs.PlateCarree()
        
        grid_dist = np.max([np.round(np.max([abs(self.extent_ll[1] - self.extent_ll[0])/3, abs(self.extent_ll[3] - self.extent_ll[2])/3]), decimals = 1), 0.1])

        gl = ax.gridlines(crs=crs_lonlat,
                          xlocs=np.arange(np.floor(self.extent_ll[0]), np.ceil(self.extent_ll[1])+.5, grid_dist),
                          ylocs=np.arange(np.floor(self.extent_ll[2]), np.ceil(self.extent_ll[3])+.5, grid_dist),
                          draw_labels=True, linestyle='-.', zorder=2)
        
        gl.xlabels_top = None
        gl.ylabels_right = None
        gl.xformatter = LONGITUDE_FORMATTER
        gl.yformatter = LATITUDE_FORMATTER
        
        data = self.Columns[varname]['TREND']
        data_p  =self.Columns[varname]['PVALUE']

        if varname == "ET":
            color_map = 'RdYlGn_r'
        else:
            color_map = 'RdYlGn'
            
        im2 = ax.imshow(data_p, extent=self.extent_ll, origin='upper', zorder = 1, cmap = 'gray_r', vmin = 0., vmax = 1.)
        
        cm = im2.get_cmap()
        cm.set_bad(alpha = 0.0)
        
        im = ax.imshow(data, extent=self.extent_ll, cmap=color_map,
                       origin='upper', zorder=0, vmin = -15, vmax = 15)
        
        current_cmap = im.get_cmap()
        current_cmap.set_bad(color='#d8dcd6')

        ax.add_feature(feature.OCEAN)
        ax.add_feature(feature.COASTLINE, linewidth=4)
        
        cntrs = shpreader.natural_earth(resolution='50m', category='cultural', name='admin_0_countries')
        country_list = [country.geometry for country in shpreader.Reader(cntrs).records() if country.attributes['CONTINENT'] == 'Africa']

        ax.add_geometries(country_list, ccrs.Geodetic(), edgecolor='k', facecolor='none', linestyle ='-', linewidth = 0.2)
                             
        divider = make_axes_locatable(ax)
        ax_cb = divider.new_vertical(size="5%", pad=0.3, axes_class=plt.Axes, pack_start = True)

        self.f.add_axes(ax_cb)
            
        cbar = plt.colorbar(im, cax=ax_cb, orientation = "horizontal", extend = 'both')
        
        cbar.set_label("Trend [%/year]")
        
        import matplotlib.patches as mpatches
        
        patches = [ mpatches.Patch(color='k', label="p-value > 0.1")]
        ax.legend(handles=patches)#, bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0. )

    def save(self):
        
        self.f.suptitle("{2} \n{0} till {1}, {4} ha @ {3} meter/pixel".format(self.Start, self.End, self.Name, int(self.scale.getInfo()), int(self.Area / 10000)), fontsize = 18)           

        self.savepath = os.path.join(self.OutFldr, self.Name + '.pdf')
        
        plt.savefig(self.savepath)
        plt.savefig(self.savepath.replace('.pdf', '.png'))
        
    def calcYIELD(self):
        
        def _calcYIELD(img):
            year = img.get('system:time_start')
            unit = img.get('unit')
            img = img.multiply(ee.Number(img.get('multiplier'))) ####
            img = img.multiply(self.HI).divide(1 - self.WC)
            img = img.set('system:time_start', year).set('multiplier', 1.0).set('unit', unit).set('time_resolution', 'YEAR')
            return img
        
        self.VARS['YIELD'] = self.VARS['AGBP'].map(_calcYIELD).set('variable', 'Yield')
    
    def fillEXCL(self, path = r"C:\Users\bec\Desktop\test\Book1_proj.xlsx"):
        
        self.layout_file = path
        wb = load_workbook(self.layout_file)
        ws = wb['Sheet1']
        
        ws['A1'] = self.Name
        ws['D17'] = self.HI
        ws['D19'] = self.WC
        
        j = 6

        for i, val in enumerate(sheet.Columns['AGBP']['TS'][1]): #kg/ha
            ws['C{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['YIELD']['TS'][1]): #kg/ha
            ws['D{0}'.format(i + j)] = val
            
        yield_abs = sheet.Columns['YIELD']['TS'][1] * (self.Area/1e4) / 1e9 #Mton
        for i, val in enumerate(yield_abs * 1000): #kton
            ws['E{0}'.format(i + j)] = val
            
        att_yield_abs = sheet.Columns['YIELD']['TS95'][1] * (self.Area/1e4) / 1e9 #Mton
        for i, val in enumerate(att_yield_abs * 1000): #kton
            ws['F{0}'.format(i + j)] = val
            
        for i, val in enumerate((att_yield_abs - yield_abs)*1000): #kton
            ws['G{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['ET']['TS'][1]): #mm
            ws['H{0}'.format(i + j)] = val        

        et_abs = sheet.Columns['ET']['TS'][1] / 1000 * self.Area / 1e9 #km3
        for i, val in enumerate(et_abs * 1000): # 0.001 km3
            ws['I{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['ET0']['TS'][1]): # mm
            ws['J{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['P']['TS'][1]): # mm
            ws['K{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['WP']['TS'][1]): # kg/m3
            ws['L{0}'.format(i + j)] = val
            
        for i, val in enumerate(sheet.Columns['WP']['TS95'][1]): # kg/m3
            ws['M{0}'.format(i + j)] = val

        cons_gap = et_abs - ((yield_abs * 1e9 / sheet.Columns['WP']['TS95'][1]) / 1e9) #km3
        for i, val in enumerate(cons_gap * 1000): # 0.001 km3
            ws['N{0}'.format(i + j)] = val
            
        out = os.path.join(self.OutFldr, "{0}_table.xlsx".format(self.Name))
        wb.save(out)
        
        o = win32com.client.Dispatch("Excel.Application")
        o.Visible = False
        
        wb = o.Workbooks.Open(out)
        
        self.path_to_pdf = out.replace('xlsx', 'pdf')
        
        wb.WorkSheets([1]).Select()
        
        wb.ActiveSheet.ExportAsFixedFormat(0, self.path_to_pdf)
        
        o.Workbooks.Close()
        
    def combine(self):
        
        from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
        
        pages_to_delete = [1, 2] # page numbering starts from 0
        infile = PdfFileReader(self.path_to_pdf, 'rb')
        output = PdfFileWriter()
        
        for i in range(infile.getNumPages()):
            if i not in pages_to_delete:
                p = infile.getPage(i)
                output.addPage(p)
        
        with open(self.path_to_pdf.replace('.pdf', '_new.pdf'), 'wb') as f:
            output.write(f)
            
        x = [self.savepath, self.path_to_pdf.replace('.pdf', '_new.pdf')]

        merger = PdfFileMerger()

        for pdf in x:
            merger.append(open(pdf, 'rb'))
        
        with open(self.path_to_pdf.replace('.pdf', '_combined.pdf'), "wb") as fout:
            merger.write(fout)
            
        os.remove(self.path_to_pdf.replace('.pdf', '_new.pdf'))
        os.remove(self.path_to_pdf)
    
    def make(self):
        
        self.calcYIELD()
        self.calcWP()

        for var in ['AGBP', 'ET', 'WP']:
            self.createTRENDMAP(var)
            self.createTS(var)
            
            self.plotHIST(var)
            self.plotTRENDMAP(var)
            self.plotTS(var)
        
        self.createTS('P')
        self.createTS('ET0')
        
        self.createTS('YIELD')
        self.createTS('YIELD', reducer = ee.Reducer.percentile([95]))
        self.createTS('WP', reducer = ee.Reducer.percentile([95]))


        self.save()
        self.fillEXCL(path = r"C:\Users\bec\Desktop\test\Book1_proj.xlsx")
        
        self.combine()
        
    def make2(self):
        
        self.calcYIELD()
        self.calcWP(base = 'YIELD')

        for var in ['YIELD', 'ET', 'WP']:
            self.createTRENDMAP(var)
            self.createTS(var)
            
            self.plotHIST(var)
            self.plotTRENDMAP(var)
            self.plotTS(var)
        
        self.createTS('P')
        self.createTS('ET0')
        
        self.createTS('AGBP')
        self.createTS('YIELD', reducer = ee.Reducer.percentile([95]))
        self.createTS('WP', reducer = ee.Reducer.percentile([95]))

        self.save()
        self.fillEXCL(path = r"C:\Users\bec\Desktop\test\Book1_proj.xlsx")
        
        self.combine()
        
#%%    


# 100m example
sheet = trend_sheet('Wonji', 
                    None, 
                    'projects/fao-wapor/L2/L2_AETI_A',
                    'projects/fao-wapor/L1/L1_PCP_E',
                    'projects/fao-wapor/L1/L1_RET_E',
                    'users/bcoerver/Ethiopia_simon/wonji', 
                    r"C:\Users\bec\Desktop\test",
                    NPP = 'projects/fao-wapor/L2/L2_NPP_D')

sheet.make()

# 250m example
sheet = trend_sheet('Githongo', 
                    'projects/fao-wapor/L1/L1_AGBP_A', 
                    'projects/fao-wapor/L1/L1_AETI_A',
                    'projects/fao-wapor/L1/L1_PCP_E',
                    'projects/fao-wapor/L1/L1_RET_E',
                    'users/bcoerver/Kenya/ThirdEye/Githongo', 
                    r"C:\Users\bec\Desktop\test",
                    NPP = None)

sheet.make()

# example with yield calculation, for now the HI and WC needs to be adjusted 
# inside the code (this should become an input and the functions "make()" and "make2()" should be merged)
sheet = trend_sheet('Mafambisse', 
                'projects/fao-wapor/L1/L1_AGBP_A', 
                'projects/fao-wapor/L1/L1_AETI_A',
                'projects/fao-wapor/L1/L1_PCP_E',
                'projects/fao-wapor/L1/L1_RET_E',
                'users/bcoerver/Mozambique/Mafambisse', 
                r"C:\Users\bec\Desktop\test")

sheet.make2()
#%%
#
#shapes = [
#          (r'users/bcoerver/Ethiopia_simon/metahara', 'Metahara - Unknown Crops'),
#          (r'users/bcoerver/Ethiopia_simon/other', 'Other - Mixed Crops'),
#          (r'users/bcoerver/Ethiopia_simon/tibila', 'Tibila - Unknown Crops'),
#          (r'users/bcoerver/Ethiopia_simon/uaaie', 'Uaaie - Tree Crops'),
#          (r'users/bcoerver/Kenya/Mara_Basin', 'Mara Basin'),
#          (r'users/bcoerver/Kenya/Mara_Kenya', 'Mara Basin (Kenya)'),
#          (r'users/bcoerver/Rwanda/Existing_Terraces', 'Rwanda Terraces'),
#          (r'users/bcoerver/Rwanda/Rwanda_NE', 'Rwanda Terraces NE'),
#          (r'users/bcoerver/Rwanda/Rwanda_NW', 'Rwanda Terraces NW'),
#          (r'users/bcoerver/Rwanda/Rwanda_SW', 'Rwanda Terraces SW'),
#          (r'users/bcoerver/gadm36_PSE_0', 'West Bank (Irrigated)'),
#          (r'users/bcoerver/Nigeria/Kaduna', 'Kaduna (Nigeria)'),
#          ('users/bcoerver/Rwanda/Bralirwa_pivots', 'Bralirwa'),
#           # ThirdEye
#          (r'users/bcoerver/Kenya/ThirdEye/Githongo', 'Githongo'),
#          (r'users/bcoerver/Kenya/ThirdEye/Kibirichia', 'Kibirichia'),
#          (r'users/bcoerver/Kenya/ThirdEye/Meru_county', 'Meru County'),
#          (r'users/bcoerver/Kenya/ThirdEye/Miathene', 'Miathene'),
#           # DryDev
#          (r'users/bcoerver/Kenya/DryDev/enziu_comparisonsite', 'Enziu (comparison)'),
#          (r'users/bcoerver/Kenya/DryDev/ikatini_interventionsite', 'Ikatini (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/ititu_interventionsite', 'Ititu (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/itunduimuni_comparisonsite', 'Itunduimuni (comparison)'),
#          (r'users/bcoerver/Kenya/DryDev/kamwala_comparisonsite', 'Kamwala (comparison)'),
#          (r'users/bcoerver/Kenya/DryDev/kathoka_interventionsite', 'Kathoka (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/kibau_interventionsite', 'Kibau (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/mukange-lower_interventionsite', 'Lower Mukange (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/mukange-upper_comparisonsite', 'Upper Mukange (comparison)'),
#          (r'users/bcoerver/Kenya/DryDev/mukange-upper_interventionsite', 'Upper Mukange (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/syomunyu_interventionsite', 'Syomunyu (intervention)'),
#          (r'users/bcoerver/Kenya/DryDev/syumakandu_comparisonsite', 'Syumakandu (comparison)'),
#          (r'users/bcoerver/Kenya/DryDev/thwake_comparisonsite', 'Thwake (comparison)'),
#           # SWA
#          (r'users/bcoerver/Kenya/SWA/Laikipia_county', 'Laikipia County'),
#          (r'users/bcoerver/Kenya/SWA/Machakos_county', 'Machakos County'),
#          (r'users/bcoerver/Kenya/SWA/Nakuru_county', 'Nakuru County'),
#          (r'users/bcoerver/Kenya/SWA/UasinGishu_county', 'Uasin Gishu County'),
#           # Burkina Faso (Not available in 100m res.)
#          (r'users/bcoerver/BurkinaFaso/B07_Comparaison_LankoueFinal_region','Lankoue (comparison)'),
#          (r'users/bcoerver/BurkinaFaso/B07_Intervention_KiembaraFinal_region','Kiembara (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B08_ComparaisonTougoFinal_region','Tougo (comparison)'),
#          (r'users/bcoerver/BurkinaFaso/B08_Interv_BassiFinal_region','Bassi (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B09_ComparaisonTangayeFinal_region','Tangaye (comparison)'),
#          (r'users/bcoerver/BurkinaFaso/B09_InterventionZogoreFinal_region','Zogore (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B13_Comparaison_RoukoFinal_region','Rouko (comparison)'),
#          (r'users/bcoerver/BurkinaFaso/B13_Intervention_KongoussiTikareFinal_region','KongoussiTikare (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B16_ComparaisonKirsiFinal_region','Kirsi (comparison)'),
#          (r'users/bcoerver/BurkinaFaso/B16_InterventionArbolleFinal_region','Arbolle (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B27_Intervention_KyonFinal_region','Kyon (intervention)'),
#          (r'users/bcoerver/BurkinaFaso/B27_Comparaison_Pouni_region','Pouni (comparison)'),
#          # Mozambique
#          (r'users/bcoerver/Mozambique/Xinavane_WGS84_dissolved','Xinavane'),
#          (r'users/bcoerver/Mozambique/Barue', 'Barue'),
#          (r'users/bcoerver/Mozambique/Manica', 'Manica'),
#          (r'users/bcoerver/Mozambique/Moatize', 'Moatize'),
#          (r'users/bcoerver/Mozambique/Nhamatanda', 'Nhamatanda'),
#          (r'users/bcoerver/Mozambique/Sem_Provincias', 'Sem_Provincias'),
#          (r'users/bcoerver/Mozambique/Sofala', 'Sofala'),
#          (r'users/bcoerver/Mozambique/Tete', 'Tete'),
#          (r'users/bcoerver/Mozambique/Zambezia', 'Zambezia'),
#          (r'users/bcoerver/Mozambique/IWACATECH/Beira', 'Beira'),
#          (r'users/bcoerver/Mozambique/IWACATECH/Manhica', 'Manhica'),
#          (r'users/bcoerver/Mozambique/IWACATECH/xinavane_centre_pivot', 'Xinavane (centre pivot)'),
#          (r'users/bcoerver/Mozambique/ARA_Zambeze_1', 'ARA Zambeze'),
#         ]
#
#shapes_sugarcane = [
#                    (r'users/bcoerver/Mozambique/Mafambisse', 'Mafambisse'),
#                    (r'users/bcoerver/Mozambique/Xinavane_1', 'Xinavane_1'),
#                    (r'users/bcoerver/Ethiopia_simon/wonji', 'Wonji - Sugarcane'),
#                   ]
