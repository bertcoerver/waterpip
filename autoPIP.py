# -*- coding: utf-8 -*-
"""
Created on Mon Sep 10 16:03:12 2018

@author: bec
"""
import ee
import os
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from WA_Hyperloop import becgis
import numpy as np  
import gdal
import cartopy.crs as ccrs
from cartopy.mpl.gridliner import LONGITUDE_FORMATTER, LATITUDE_FORMATTER
import cartopy.feature as feature
from mpl_toolkits.axes_grid1 import make_axes_locatable
import cartopy.io.shapereader as shpreader
from openpyxl import load_workbook
import scipy.stats as sts
import win32com.client
import geegis

ee.Initialize()

class trend_sheet(object):
    
    def __init__(self, name, shape, OutputFolder, mask = False):
        
        AGBP = 'projects/fao-wapor/L2/L2_AGBP_S'
        WP = 'projects/fao-wapor/L2/L2_GBWP_S'
        P = 'projects/fao-wapor/L1/L1_PCP_E'
        ETref = 'projects/fao-wapor/L1/L1_RET_E'
        PHEN = 'projects/fao-wapor/L2/L2_PHE_S'
        LULC = 'projects/fao-wapor/L1/L1_LCC_A'

        # Turn of the visibility of matplotlib figures (i.e. figures remain invisble until saved as a file or after plt.show())
        plt.ioff()
        
        self.Name = name
        
        # Create figure object
        self.initFigure()
        
        # Create a mask image from the shapefile
        self.CountryShape = ee.FeatureCollection(shape)
        self.CountryMask = self.CountryShape.map(lambda sh: sh.set('1', 1)).reduceToImage(['1'], ee.Reducer.max()).unmask().cast({"max": "uint8"})

        self.VARS = {"WP": ee.ImageCollection(WP).set('variable', 'Water Productivity'),
                     "P": ee.ImageCollection(P).set('variable', 'Precipitation'),
                     "AGBP": ee.ImageCollection(AGBP).set('variable', 'Above Ground Biomass Prod.'),
                     "ET0": ee.ImageCollection(ETref).set('variable', "Reference ET"),
                     "PHEN": ee.ImageCollection(PHEN).set('variable', "Phenology")}
        
        self.Start = ee.Date(self.VARS['WP'].aggregate_min('system:time_start')).get('year').getInfo()
        self.End = ee.Date(self.VARS['WP'].aggregate_max('system:time_start')).get('year').getInfo()
          
        if mask:
            lulc = ee.ImageCollection(LULC).filterDate('2013-01-01', '2013-12-31').first()
            lulc = ee.Image(lulc)
            def _mask_LULC(img):
                img = ee.Image(img)
                if len(mask) == 1:
                    img_masked = img.updateMask(lulc.eq(mask[0]))
                elif len(mask) == 2:
                    img_masked = img.updateMask(lulc.gte(41)).updateMask(lulc.lt(43))
                return ee.Image(img_masked).copyProperties(img, img.propertyNames())
            
            for varname, VAR in self.VARS.items():
                self.VARS[varname] = ee.ImageCollection(VAR.map(_mask_LULC).copyProperties(VAR, VAR.propertyNames()))
                
        self.correct_metadata_AGBP()
        
        self.VARS["P"] = self.make_seasons("P")
        self.VARS["ET0"] = self.make_seasons("ET0")
        
        def _mask_nodata(img):
            img = ee.Image(img)
            return img.updateMask(img.gte(0))
        
        def _apply_multiplier(img):
            return img.multiply(ee.Number(img.get("multiplier"))).copyProperties(img,img.propertyNames().remove('multiplier'))
        
        for varname, VAR in self.VARS.items():
            self.VARS[varname] = ee.ImageCollection(VAR.map(_mask_nodata))
            
        for varname, VAR in self.VARS.items():
            print("applying multiplier to", varname)
            self.VARS[varname] = ee.ImageCollection(VAR.map(_apply_multiplier))
            
        def _calc_area(img):
            return ee.Image.pixelArea().multiply(img.gte(0)).copyProperties(img, img.propertyNames()).set('unit', 'm2')
        
        self.VARS['areas'] = self.VARS['WP'].map(_calc_area).select([0], ['b1'])
        
        self.calcET()
        
        # Determine the resolution to be used
        img = ee.Image(self.VARS['AGBP'].first())
        self.scale = img.select(img.bandNames()).projection().nominalScale()

        # Calculate the total area of the AOI
        self.Area = self.CountryShape.geometry().area().getInfo()

        self.OutFldr = OutputFolder
        
        # Set which variables to show in which columns
        self.Columns = {'AGBP_yr': {'pos': 0}, 'ET_yr': {'pos': 1}, 'WP_yr': {'pos': 2}}

    def initFigure(self):
        self.f = plt.figure(1)
        self.f.clf()
        self.f.set_dpi(96)
        self.f.set_size_inches(16.54, 11.69) # A3 size
        self.gs = gridspec.GridSpec(3, 3, height_ratios=[1,1,2])
        
    def correct_metadata_AGBP(self):
        def _correct_AGBP(img):
            new_img = img.set('multiplier', 1.0)
            return new_img
        
        self.VARS['AGBP'] = self.VARS['AGBP'].map(_correct_AGBP)
        
    def calcWP(self, base = 'AGBP_yr'):
        
        def _calcWP(year):
            agbp = ee.Image(self.VARS[base].filterMetadata('system:time_start', 'equals', year).first())
            et = ee.Image(self.VARS['ET_yr'].filterMetadata('system:time_start', 'equals', year).first())
            wp = agbp.divide(et.multiply(10)).set('system:time_start', year).set('unit', 'kg/m3')
            
            return wp
    
        years = ee.List(self.VARS[base].aggregate_array('system:time_start'))
        self.VARS['WP_yr'] = ee.ImageCollection(years.map(_calcWP)).set('variable', 'Water Productivity')

    def calcET(self):
        
        def _calcET(agbp):
            
            current_year = agbp.get('system:time_start')
            current_season = agbp.get('season')
            
            wp = ee.Image(self.VARS['WP'].filterMetadata('system:time_start', 'equals', current_year).filterMetadata('season', 'equals', current_season).first())

            et = agbp.divide(wp).divide(10) # mm
            
            et = et.set('system:time_start', current_year).set('season', current_season).set('unit', 'mm') 
            
            return et

        self.VARS['ET'] = ee.ImageCollection(self.VARS['AGBP'].map(_calcET)).set('variable', 'ET')

    def calcTREND(self, VAR):
        
        band_name = ee.Image(VAR.first()).bandNames().getInfo()[0]
        
        start = VAR.aggregate_min('system:time_start')
        first = ee.Image(VAR.filterMetadata('system:time_start', 'equals', start).first()).select(band_name)
          
        trend = VAR.select(band_name).formaTrend()
        
        def add_timeband(VARimg):
            date = ee.Date(VARimg.get("system:time_start"))
            return ee.Image(date.get("year")).cast({"constant": "int32"}).addBands(VARimg)
      
        with_time = VAR.map(add_timeband);
        pearson = with_time.reduce(ee.Reducer.pearsonsCorrelation());
      
        trend = trend.select('long-trend').divide(first).multiply(100)
        trend = trend.addBands(self.CountryMask).addBands(pearson)
        
        return trend 
            
    def createTRENDMAP(self, varname):
        trend = self.calcTREND(self.VARS[varname])
        output_fh = os.path.join(self.OutFldr, '{0}.zip'.format(varname))
        
        self.factor = np.interp(self.Area, [50000, 625000000], [0.1, 1.0])
        scale = self.scale.multiply(self.factor).getInfo()

        geegis.downloadImage(trend, output_fh, self.CountryShape, scale)
        
        TREND = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.long-trend.tif'))
        MASK = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.max.tif'))
        PVALUE = becgis.OpenAsArray(os.path.join(output_fh[:-4], 'test.p-value.tif'))
        
        assert np.shape(TREND) == np.shape(MASK), "resolution dont match"
        
        TREND[MASK == 0] = np.nan
        TREND[TREND == 0.0000000000] = np.nan # need to fix this properly.
        TREND[PVALUE > 0.1] = np.nan

        PVALUE[PVALUE <= 0.1] = np.nan
        PVALUE[MASK == 0] = np.nan
        PVALUE[~np.isnan(PVALUE)] = 1.
        
        AREA = becgis.MapPixelAreakm(os.path.join(output_fh[:-4], 'test.max.tif'))
        self.PixelArea = np.mean(AREA[MASK == 1]) * 100 # ha
        
        self.Columns[varname]['TREND'] = TREND
        self.Columns[varname]['PVALUE'] = PVALUE
        
        ds = gdal.Open(os.path.join(output_fh[:-4], 'test.long-trend.tif'))
        gt = ds.GetGeoTransform()
        n_cols = ds.RasterXSize
        n_rows = ds.RasterYSize
        ds = None
        self.extent_ll = (gt[0], gt[0] + (gt[1] * n_cols), gt[3] + (gt[5] * n_rows), gt[3])
    
    def plotTS(self, varname):
        
        ax = plt.subplot(self.gs[0, self.Columns[varname]['pos']])
            
        yrs = np.array(ee.List(self.TS[varname]['system:time_start'].tolist()).map(lambda x: ee.Date(x).get('year')).getInfo())
        vals = self.TS[varname]['b1_mean']
        unit = '[' + str(np.unique(self.TS[varname]['unit'])[0]) + ']'
    
        ax.set_ylabel("{0} {1}".format(varname, unit.replace('/yr', '')))
        
        slope, intercept, r_value, p_value, std_err = sts.linregress(yrs, vals)
        
        fit = np.polyfit(yrs, vals, 1)
        fit_fn = np.poly1d(fit)
        
        yrs_long = [yrs[0]-1, yrs[-1]+1]
    
        ax.plot(yrs_long, fit_fn(yrs_long), 'k--')

        ax.set_ylim([0.93 * np.min(vals), 1.02 * np.max(vals)])
        ax.set_xlim(yrs_long)
        ax.set_title("{4} \n slope = {0:.2f} {1}, p = {2:.3f}, r = {3:.2f}".format(slope, unit[1:-1], p_value, r_value, varname))
        ax.set_facecolor('#d8dcd6')
        ax.grid(b=True, which='Major', color='0.65',linestyle='--', zorder=0)
        ax.bar(yrs, vals, color = '#005c30', zorder = 2)
        ax.set_xticks(yrs)
        ax.set_xticklabels(['`' + yr[-2:] for yr in yrs.astype(str)])
        
    def plotHIST(self, varname):
        ax = plt.subplot(self.gs[1, self.Columns[varname]['pos']])
        ax.set_xlabel("Trend [%/yr]")
        ax.set_ylabel("Area [ha]")
        ax.set_facecolor('#d8dcd6')
        ax.grid(b=True, which='Major', color='0.65',linestyle='--', zorder=0)

        rng = (np.floor(np.nanmin(self.Columns[varname]['TREND']) / 10)*10, np.ceil(np.nanmax(self.Columns[varname]['TREND']) / 10)*10)
        
        try:
            bns = np.min([int(np.diff(rng)[0]), 21])
        except:
            bns = 0
        
        print(varname, bns, rng)
        
        if bns != 0:
            pixels, bins = np.histogram(self.Columns[varname]['TREND'][~np.isnan(self.Columns[varname]['TREND'])], range = rng, bins = bns)
    
            bins = (bins + np.diff(bins)[0]/2)[:-1]
            
            wdt = np.diff(bins)[0]
    
            if varname == "ET":
                colors = ['#005c30','#940121']
            else:
                colors = ['#940121','#005c30']
                                  
            ax.bar(bins[bins < 0], pixels[bins < 0], color = colors[0], width = wdt, zorder = 2)
            ax.bar(bins[bins >= 0], pixels[bins >= 0], color = colors[1], width = wdt, zorder = 2)
            
        new_lbls = [str(int(item * self.PixelArea)) for item in ax.get_yticks().tolist()]
        
        ax.set_yticklabels(new_lbls)
        
    def plotTRENDMAP(self, varname):
        ax = plt.subplot(self.gs[2, self.Columns[varname]['pos']], projection = ccrs.PlateCarree())
        
        crs_lonlat = ccrs.PlateCarree()
        
        grid_dist = np.max([np.round(np.max([abs(self.extent_ll[1] - self.extent_ll[0])/3, abs(self.extent_ll[3] - self.extent_ll[2])/3]), decimals = 1), 0.1])

        gl = ax.gridlines(crs=crs_lonlat,
                          xlocs=np.arange(np.floor(self.extent_ll[0]), np.ceil(self.extent_ll[1])+.5, grid_dist),
                          ylocs=np.arange(np.floor(self.extent_ll[2]), np.ceil(self.extent_ll[3])+.5, grid_dist),
                          draw_labels=True, linestyle='-.', zorder=2)
        
        gl.xlabels_top = None
        gl.ylabels_right = None
        gl.xformatter = LONGITUDE_FORMATTER
        gl.yformatter = LATITUDE_FORMATTER
        
        data = self.Columns[varname]['TREND']
        data_p  =self.Columns[varname]['PVALUE']

        if varname == "ET":
            color_map = 'RdYlGn_r'
        else:
            color_map = 'RdYlGn'
            
        im2 = ax.imshow(data_p, extent=self.extent_ll, origin='upper', zorder = 1, cmap = 'gray_r', vmin = 0., vmax = 1.)
        
        cm = im2.get_cmap()
        cm.set_bad(alpha = 0.0)
        
        im = ax.imshow(data, extent=self.extent_ll, cmap=color_map,
                       origin='upper', zorder=0, vmin = -15, vmax = 15)
        
        current_cmap = im.get_cmap()
        current_cmap.set_bad(color='#d8dcd6')

        ax.add_feature(feature.OCEAN)
        ax.add_feature(feature.COASTLINE, linewidth=4)
        
        cntrs = shpreader.natural_earth(resolution='50m', category='cultural', name='admin_0_countries')
        country_list = [country.geometry for country in shpreader.Reader(cntrs).records() if country.attributes['CONTINENT'] == 'Africa']

        ax.add_geometries(country_list, ccrs.Geodetic(), edgecolor='k', facecolor='none', linestyle ='-', linewidth = 0.2)
                             
        divider = make_axes_locatable(ax)
        ax_cb = divider.new_vertical(size="5%", pad=0.3, axes_class=plt.Axes, pack_start = True)

        self.f.add_axes(ax_cb)
            
        cbar = plt.colorbar(im, cax=ax_cb, orientation = "horizontal", extend = 'both')
        
        cbar.set_label("Trend [%/year]")
        
        import matplotlib.patches as mpatches

        patches = [ mpatches.Patch(color='k', label="p-value > 0.1")]
        ax.legend(handles=patches)

        
    def save(self):
        
        self.f.suptitle("{2} \n{0} till {1}, {4} ha @ {3} meter/pixel".format(self.Start, self.End, self.Name, int(self.scale.getInfo()), int(self.Area / 10000)), fontsize = 18)           

        self.savepath = os.path.join(self.OutFldr, self.Name + '.pdf')
        
        plt.savefig(self.savepath)
        plt.savefig(self.savepath.replace('.pdf', '.png'))
    
    def fillEXCL(self, path = r"D:\Github\waterpip\Book1_proj_detailed.xlsx", years_only = False):
        
        self.layout_file = path
        wb = load_workbook(self.layout_file)
        ws = wb['Sheet1']
        
        ws['A1'] = self.Name
        
        def _merge_lists(yr, sns):
            return np.array([yr, sns.reshape((9,2))[:,0], sns.reshape((9,2))[:,1]]).reshape((27), order = 'F')

        def _fill_vals(vals, column, j = 6):
            for i, val in enumerate(vals):
                ws['{0}{1}'.format(column, i + j)] = val

        # AREA [ha]
        yr_area = self.TS['areas_yr']['b1_sum'].astype(np.float64) / 1e4
        sns_area = self.TS['areas']['b1_sum'].astype(np.float64) / 1e4
        vals_area = _merge_lists(yr_area, sns_area)
        _fill_vals(vals_area, 'C')
        
        # AGBP [kg/ha]
        sns_agbp = self.TS['AGBP']['b1_mean'].astype(np.float64)
        yr_agbp = self.TS['AGBP_yr']['b1_mean'].astype(np.float64)
        vals_agbp = _merge_lists(yr_agbp, sns_agbp)
        _fill_vals(vals_agbp, 'D')
        
        # AGBP [kton]
        vals_agbp_abs = vals_agbp * vals_area / 1e6
        _fill_vals(vals_agbp_abs, 'E')

        # Attainable AGBP [kton]   
        sns_agbp_att = self.TS['AGBP']['b1_p95'].astype(np.float64) * sns_area / 1e6
        yr_agbp_att = self.TS['AGBP_yr']['b1_p95'].astype(np.float64) * yr_area / 1e6
        vals_agbp_att = _merge_lists(yr_agbp_att, sns_agbp_att)
        _fill_vals(vals_agbp_att, 'F')
        
        # AGBP Gap [kton]
        vals_agbp_gap = vals_agbp_att - vals_agbp_abs
        _fill_vals(vals_agbp_gap, 'G')
        
        # ET [mm]
        sns_et = self.TS['ET']['b1_mean'].astype(np.float64)
        yr_et = self.TS['ET_yr']['b1_mean'].astype(np.float64)
        vals_et = _merge_lists(yr_et, sns_et)
        _fill_vals(vals_et, 'H')

        # ET [0.001 km3]
        vals_et_abs = vals_et / 1e3 * vals_area * 1e4 / 1e6
        _fill_vals(vals_et_abs, 'I')
        
        # ET0 [mm]
        yr_et0 = self.TS['ET0_yr']['b1_mean'].astype(np.float64)
        sns_et0 = self.TS['ET0']['b1_mean'].astype(np.float64)
        vals_et0 = _merge_lists(yr_et0, sns_et0)
        _fill_vals(vals_et0, 'J')
        
        # P [mm]
        yr_p = self.TS['P_yr']['b1_mean'].astype(np.float64)
        sns_p = self.TS['P']['b1_mean'].astype(np.float64)
        vals_p = _merge_lists(yr_p, sns_p)
        _fill_vals(vals_p, 'K')
        
        # WP [kg/m3]
        yr_wp = self.TS['WP_yr']['b1_mean'].astype(np.float64)
        sns_wp = self.TS['WP']['b1_mean'].astype(np.float64)
        vals_wp = _merge_lists(yr_wp, sns_wp)
        _fill_vals(vals_wp, 'L')

        # Attainable WP [kg/m3]
        yr_wp_att = self.TS['WP_yr']['b1_p95'].astype(np.float64)
        sns_wp_att = self.TS['WP']['b1_p95'].astype(np.float64)
        vals_wp_att = _merge_lists(yr_wp_att, sns_wp_att)
        _fill_vals(vals_wp_att, 'M')

        # Potential ET Savings [0.001 km3] 
        vals_cons_gap = vals_et_abs - (vals_agbp_abs / vals_wp_att)
        _fill_vals(vals_cons_gap, 'N')
        
        if years_only:
            rows = [x for x in np.arange(6,33,1) if x not in np.arange(6,33,3)]
            for row in rows:
                ws.row_dimensions[row].hidden= True
        
        out = os.path.join(self.OutFldr, "{0}_table.xlsx".format(self.Name))
        wb.save(out)
        
        o = win32com.client.Dispatch("Excel.Application")
        o.Visible = False
        
        wb = o.Workbooks.Open(out)
        
        self.path_to_pdf = out.replace('xlsx', 'pdf')
        
        wb.WorkSheets([1]).Select()
        
        wb.ActiveSheet.ExportAsFixedFormat(0, self.path_to_pdf)
        
        o.Workbooks.Close()
        
    def combine(self):
        
        from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
        
        pages_to_delete = [1, 2] # page numbering starts from 0
        infile = PdfFileReader(self.path_to_pdf, 'rb')
        output = PdfFileWriter()
        
        for i in range(infile.getNumPages()):
            if i not in pages_to_delete:
                p = infile.getPage(i)
                output.addPage(p)
        
        with open(self.path_to_pdf.replace('.pdf', '_new.pdf'), 'wb') as f:
            output.write(f)
            
        x = [self.savepath, self.path_to_pdf.replace('.pdf', '_new.pdf')]

        merger = PdfFileMerger()

        for pdf in x:
            merger.append(open(pdf, 'rb'))
        
        with open(self.path_to_pdf.replace('.pdf', '_combined.pdf'), "wb") as fout:
            merger.write(fout)
            
        os.remove(self.path_to_pdf.replace('.pdf', '_new.pdf'))
        os.remove(self.path_to_pdf)
    
    def make(self):

        self.TS = dict()
        
        self.TS['areas'] = geegis.createTS(self.VARS['areas'], self.CountryShape, self.scale, ['sum'], copy_props = ['system:time_start', 'unit'])
        self.TS['ET']     = geegis.createTS(self.VARS['ET'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])
        self.TS['AGBP']   = geegis.createTS(self.VARS['AGBP'], self.CountryShape, self.scale, ['mean', 'p95'], copy_props = ['system:time_start', 'unit'])
        self.TS['WP']  = geegis.createTS(self.VARS['WP'], self.CountryShape, self.scale, ['mean', 'p95'], copy_props = ['system:time_start', 'unit'])
        self.TS['P']      = geegis.createTS(self.VARS['P'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])
        self.TS['ET0']     = geegis.createTS(self.VARS['ET0'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])

        self.VARS['areas_yr'] = self.VARS['areas']
        self.VARS['AGBP_yr'] = self.VARS['AGBP']
        self.VARS['ET_yr'] = self.VARS['ET']
        self.VARS['P_yr'] = self.VARS['P']
        self.VARS['ET0_yr'] = self.VARS['ET0']

        print("Converting to yearly...")
        # Convert time resolution to yearly for required datasets
        for varname in ['areas_yr', 'AGBP_yr', 'ET_yr', 'P_yr', 'ET0_yr']:
            VAR = self.VARS[varname]
            if varname == 'areas_yr':
                self.VARS[varname] = geegis.convert_to_yearly(VAR, [self.Start, self.End], reducer = ee.Reducer.median())
            else:
                self.VARS[varname] = geegis.convert_to_yearly(VAR, [self.Start, self.End])

        self.calcWP()

        self.TS['areas_yr'] = geegis.createTS(self.VARS['areas_yr'], self.CountryShape, self.scale, ['sum'], copy_props = ['system:time_start', 'unit'])
        self.TS['ET_yr'] = geegis.createTS(self.VARS['ET_yr'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])
        self.TS['AGBP_yr'] = geegis.createTS(self.VARS['AGBP_yr'], self.CountryShape, self.scale, ['mean', 'p95'], copy_props = ['system:time_start', 'unit'])
        self.TS['WP_yr'] = geegis.createTS(self.VARS['WP_yr'], self.CountryShape, self.scale, ['mean', 'p95'], copy_props = ['system:time_start', 'unit'])
        self.TS['P_yr'] = geegis.createTS(self.VARS['P_yr'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])
        self.TS['ET0_yr'] = geegis.createTS(self.VARS['ET0_yr'], self.CountryShape, self.scale, ['mean'], copy_props = ['system:time_start', 'unit'])
        
        for var in ['AGBP_yr', 'ET_yr', 'WP_yr']:
            self.createTRENDMAP(var)

            self.plotHIST(var)
            self.plotTRENDMAP(var)
            self.plotTS(var)

        self.save()
        self.fillEXCL(years_only = False)

        self.combine()

    
    def make_seasons(self, varname):
        
        PHENie = self.VARS['PHEN']
        COL = self.VARS[varname]
        
        def _make_seasons(string):
            start_string = ee.String(string).replace('X', 's');
            end_string = ee.String(string).replace('X', 'e');
              
            start = ee.Image(PHENie.filterMetadata('code', 'equals', start_string).first());
            end = ee.Image(PHENie.filterMetadata('code', 'equals', end_string).first());
            
            year = ee.Number.parse(ee.String(string).slice(7, 9)).add(2000);
            season = ee.Number.parse(ee.String(string).slice(10, 11))
            
            date1 = ee.Date.fromYMD(year.subtract(1), 1, 1);
            date2 = date1.advance(10, 'day');
              
            delta = date2.millis().subtract(date1.millis());
              
            start_idx = ee.Image(date1.millis()).add(start.subtract(1).multiply(delta));
            end_idx = ee.Image(date1.millis()).add(end.subtract(1).multiply(delta));
              
            def _mask_precip(img):
                current_date = ee.Image(ee.Number(img.get('system:time_start')));
                    
                masked_p = ee.Image(img).updateMask(current_date.gt(start_idx)).updateMask(current_date.lt(end_idx));

                return masked_p
            
            p_season = ee.ImageCollection(COL.map(_mask_precip)).sum().set('code', start_string).set('year', year).set('season', season).set("unit", "mm").set("system:time_start", ee.Date.fromYMD(year, 1, 1).millis()).set('multiplier', 0.1)
            
            return p_season

        seasons = ee.List(['L2_PHE_09s1_X', 'L2_PHE_09s2_X','L2_PHE_10s1_X',
                   'L2_PHE_10s2_X','L2_PHE_11s1_X','L2_PHE_11s2_X',
                   'L2_PHE_12s1_X','L2_PHE_12s2_X','L2_PHE_13s1_X',
                   'L2_PHE_13s2_X', 'L2_PHE_14s1_X','L2_PHE_14s2_X',
                   'L2_PHE_15s1_X','L2_PHE_15s2_X', 'L2_PHE_16s1_X',
                   'L2_PHE_16s2_X', 'L2_PHE_17s1_X', 'L2_PHE_17s2_X']);
        
        return ee.ImageCollection(seasons.map(_make_seasons))
        
name = 'Bralirwa'
shape = 'users/bcoerver/Rwanda/Bralirwa_pivots' 
OutputFolder = r'D:\project_WaterPIP\waterpip_1' 

# 100m example
sheet = trend_sheet(name, shape, OutputFolder)
sheet.make()

